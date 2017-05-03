import os, yaml, re
from openpyxl import load_workbook
__scirpt_path__ = os.path.dirname(os.path.abspath(__file__))
__project_path__ = os.path.dirname(os.path.dirname(__scirpt_path__))
__assets_path__ = os.path.join(__project_path__, 'Assets')

pathConfigName = os.path.join(__scirpt_path__, '../pathvars.yml')

with open(pathConfigName, 'r') as stream:
	content = yaml.load(stream)
localeDir=content['localeDir']
xlsxConfigFolder=content['xlsxConfigDir']
pattern = re.compile(r'Value:(\w+)<unicode>')

def parseOneXlsxFile(xlsxfile):
	wb = load_workbook(filename = xlsxfile)
	ws = wb.active
	idCol = ws['A']
	ids = [ cell.value for cell in idCol ]
	ids = ids[1:]
	check = set()
	for id in ids:
		if id in check:
			raise Exception('table %s has the same id %s' % (xlsxfile, id.encode('utf-8')))
		check.add(id)
	data = {}
	for col in ws.iter_cols(min_row=1, min_col=2):
		lang = col[0].value
		match = pattern.match(lang)
		if match:
			lang = match.group(1)
		item = dict(zip(ids, [ cell.value or '' for cell in col[1:] ]))
		data[lang] = item
	return data

dictionary = {}
for parent, dirs, files in os.walk(os.path.join(xlsxConfigFolder, 'locale')):
	for filename in files:
		pathname = os.path.join(parent, filename)
		if os.path.isfile(pathname) and pathname.endswith('.xlsx') and '~' not in pathname:
			data = parseOneXlsxFile(pathname)
			for lang, worddict in data.items():
				l = dictionary.setdefault(lang, {})
				for new_word_id, trans in worddict.items():
					if new_word_id in l:
						raise Exception('%s has the same word id of %s' %(pathname, new_word_id.encode('utf-8')))
					l[new_word_id] = trans

for lang, worddict in dictionary.items():
	textFilename = os.sep.join([localeDir, lang, 'dictionary', 'words.txt'])
	dir = os.path.dirname(textFilename)
	if not os.path.exists(dir):
		os.makedirs(dir)

	with open(textFilename, 'w') as f:
		for word_id, trans in sorted(worddict.items()):
			f.write('%s=%s\n' % (word_id.encode('utf-8'), trans.encode('utf-8')))
